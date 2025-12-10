import os
import io
import uuid
import tempfile
from fastapi import UploadFile, HTTPException
from main import app, storage
from openpyxl import load_workbook
from fpdf import FPDF


@app.post("/excel-to-pdf")
async def excel_to_pdf(file: UploadFile):
    """Convert an uploaded XLSX file to PDF and return a Firebase URL."""
    try:
        excel_bytes = await file.read()
        wb = load_workbook(filename=io.BytesIO(excel_bytes), data_only=True)

        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.set_font("Arial", size=12)

        for sheet in wb.worksheets:
            pdf.add_page()
            pdf.multi_cell(0, 10, f"Sheet: {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                line = " | ".join([str(cell) if cell is not None else "" for cell in row])
                pdf.multi_cell(0, 10, line)

        temp_dir = tempfile.gettempdir()
        filename = f"converted_{uuid.uuid4().hex}.pdf"
        filepath = os.path.join(temp_dir, filename)
        pdf.output(filepath)

        bucket = storage.bucket()
        blob = bucket.blob(f"converted_pdfs/{filename}")
        blob.upload_from_filename(filepath)
        blob.make_public()

        return {"status": "success", "file_url": blob.public_url}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))



